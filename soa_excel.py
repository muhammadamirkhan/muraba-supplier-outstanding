"""Export a Statement of Account as .xlsx, formatted like the client's workbook.

The download is served by Streamlit rather than the dashboard itself: the page
renders inside a sandboxed iframe, which blocks downloads a link inside it would
start, so `st.download_button` is the only route that reliably produces a file.

Cell formats, column widths and the layout are taken from the first sheet of
Zetas Zemin_ SOA.xlsx so the export opens looking like the statement Finance
already circulates.
"""
import datetime as _dt
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Formats copied from the client's workbook
DATE_FMT = "[$-409]d\\-mmm\\-yy;@"
MONEY_FMT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
PCT_FMT = "0.00%"

# Column widths from the workbook (A..H)
WIDTHS = [14.4, 13.3, 17.7, 13.0, 13.9, 12.6, 88.1, 13.0]

MONTHS = {m: i + 1 for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}

NUM_COLS = (2, 4, 5)   # invoice amount, payment amount, balance (0-indexed in the row)
DATE_COLS = (1, 3)     # invoice date, payment date

_HEAD_FILL = PatternFill("solid", fgColor="DCE6F1")
_TOT_FILL = PatternFill("solid", fgColor="F2F6FC")
_THIN = Side(style="thin", color="BFCBDB")


def _parse(d):
    """'27-Feb-25' -> date. These strings are produced by transform._fmt_date_soa,
    so the shape is ours and the round-trip is exact."""
    if not d or not isinstance(d, str):
        return None
    try:
        day, mon, yr = d.split("-")
        return _dt.date(2000 + int(yr), MONTHS[mon], int(day))
    except (ValueError, KeyError):
        return None


def build(name, ledger, asof_text, category=""):
    """-> bytes of a one-sheet .xlsx statement."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SOA"

    ws["A1"] = name
    ws["A1"].font = Font(bold=True, size=13)
    ws["H1"] = asof_text
    ws["H1"].alignment = Alignment(horizontal="right")
    ws["A2"] = f"Statement of Account as of {asof_text}"
    ws["A2"].font = Font(italic=True, color="6B7686")
    if category:
        ws["A3"] = category
        ws["A3"].font = Font(size=9, color="6B7686")

    head = ledger.get("head", [])
    for c, title in enumerate(head, start=1):
        cell = ws.cell(row=4, column=c, value=title)
        cell.font = Font(bold=True)
        cell.fill = _HEAD_FILL
        cell.border = Border(bottom=_THIN)
        cell.alignment = Alignment(horizontal="right" if c - 1 in NUM_COLS else "left")

    r = 5
    for row in ledger.get("rows", []):
        for i, val in enumerate(row):
            cell = ws.cell(row=r, column=i + 1)
            if i in DATE_COLS:
                d = _parse(val)
                if d:
                    cell.value = d
                    cell.number_format = DATE_FMT
                else:
                    cell.value = val or None
            elif i in NUM_COLS:
                cell.value = val if isinstance(val, (int, float)) else None
                cell.number_format = MONEY_FMT
            else:
                cell.value = val or None
        r += 1

    cur = ledger.get("currency", "AED")
    inv, paid = ledger.get("invoiced", 0), ledger.get("paid", 0)

    tot = r
    ws.cell(row=tot, column=1, value=f"Total - {cur}")
    ws.cell(row=tot, column=3, value=inv).number_format = MONEY_FMT
    ws.cell(row=tot, column=5, value=paid).number_format = MONEY_FMT
    ws.cell(row=tot, column=6, value=inv - paid).number_format = MONEY_FMT
    ws.cell(row=tot, column=8, value="-")
    for c in range(1, 9):
        cell = ws.cell(row=tot, column=c)
        cell.font = Font(bold=True)
        cell.fill = _TOT_FILL
        cell.border = Border(top=Side(style="thin", color="4472C4"))

    pct = tot + 1
    ws.cell(row=pct, column=1, value="% of Amount paid").font = Font(bold=True)
    ws.cell(row=pct, column=3, value=(paid / inv) if inv else 0).number_format = PCT_FMT
    ws.cell(row=pct, column=3).font = Font(bold=True)

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
